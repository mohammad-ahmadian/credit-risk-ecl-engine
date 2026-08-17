import sys
import os

# Ensure root folder is in python path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '../../')))

import numpy as np
import pandas as pd
from datetime import datetime
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from src.database.db_connection import get_db_engine

class IFRS9ExpectedCreditLossEngine:
    """
    IFRS 9 Expected Credit Loss (ECL) & Staging Engine:
    - Computes LGD (Loss Given Default) with collateral haircuts
    - Computes EAD (Exposure at Default)
    - Classifies loans into IFRS 9 Stage 1, Stage 2 (SICR), or Stage 3 (Default)
    - Calculates Stage-specific ECL Provisions
    - Stores results in PostgreSQL
    - Exports formatted Auditable Executive Excel Workbook (.xlsx)
    """

    def __init__(self, collateral_haircut=0.20):
        self.engine = get_db_engine()
        self.haircut = collateral_haircut

    def fetch_credit_data(self):
        """Queries loan portfolio data, borrower info, and PD predictions from DB."""
        query = """
            SELECT 
                l.loan_id,
                l.loan_code,
                b.borrower_code,
                b.annual_income,
                b.credit_score_bureau,
                l.loan_amount,
                l.current_balance,
                l.collateral_value,
                l.days_past_due,
                l.is_defaulted,
                p.pd_12m,
                p.pd_lifetime,
                p.scorecard_points
            FROM fact_loans l
            JOIN dim_borrowers b ON l.borrower_id = b.borrower_id
            JOIN pd_model_predictions p ON b.borrower_code = p.borrower_code;
        """
        with self.engine.connect() as conn:
            df = pd.read_sql(query, conn)
        print(f"✅ Loaded {len(df)} loans with PD model estimates from PostgreSQL.")
        return df

    def calculate_ifrs9_ecl(self):
        """Computes LGD, EAD, IFRS 9 Staging, and ECL provisions."""
        df = self.fetch_credit_data()

        # 1. Calculate EAD (Exposure at Default = Current Carrying Balance)
        df['ead'] = df['current_balance']

        # 2. Calculate LGD (Loss Given Default accounting for collateral haircuts)
        net_collateral = df['collateral_value'] * (1.0 - self.haircut)
        uncovered_exposure = np.maximum(0.0, df['ead'] - net_collateral)
        df['lgd'] = np.clip(uncovered_exposure / df['ead'], 0.05, 0.90)

        # 3. IFRS 9 Staging Classification Logic
        stages = []
        ecl_amounts = []

        for idx, row in df.iterrows():
            dpd = row['days_past_due']
            pd12 = row['pd_12m']
            pd_lt = row['pd_lifetime']
            lgd = row['lgd']
            ead = row['ead']
            is_def = row['is_defaulted']

            if is_def == 1 or dpd >= 90:
                stage = 3  # Credit-Impaired / Defaulted -> Full Exposure Loss
                ecl = 1.0 * lgd * ead
            elif dpd >= 30 or pd12 >= 0.10:
                stage = 2  # SICR -> Lifetime ECL
                ecl = pd_lt * lgd * ead
            else:
                stage = 1  # Performing -> 12-Month ECL
                ecl = pd12 * lgd * ead

            stages.append(int(stage))
            ecl_amounts.append(round(ecl, 2))

        df['ifrs9_stage'] = stages
        df['ecl_amount'] = ecl_amounts
        df['calc_date'] = datetime.now().strftime('%Y-%m-%d')

        print("\n=======================================================")
        print("📊 IFRS 9 STAGING & ECL PROVISIONING SUMMARY")
        print("=======================================================")
        summary = df.groupby('ifrs9_stage').agg(
            loan_count=('loan_id', 'count'),
            total_ead=('ead', 'sum'),
            total_ecl=('ecl_amount', 'sum'),
            avg_pd_12m=('pd_12m', 'mean'),
            avg_lgd=('lgd', 'mean')
        ).reset_index()

        summary['coverage_ratio_pct'] = (summary['total_ecl'] / summary['total_ead']) * 100
        print(summary.to_string(index=False))
        print("=======================================================\n")

        # 4. Save ECL Results to PostgreSQL
        self._save_to_postgresql(df)

        # 5. Export Auditable Financial Excel Workbook
        self._export_excel_workbook(df, summary)

    def _save_to_postgresql(self, df):
        """Saves loan-level ECL results to fact_ecl_impairment table."""
        df_ecl = df[['loan_id', 'calc_date', 'ifrs9_stage', 'pd_12m', 'pd_lifetime', 'lgd', 'ead', 'ecl_amount']].copy()
        df_ecl['calc_date'] = pd.to_datetime(df_ecl['calc_date']).dt.date

        print(f"📥 Uploading {len(df_ecl)} ECL impairment records to PostgreSQL...")
        with self.engine.begin() as conn:
            conn.execute(text("TRUNCATE TABLE fact_ecl_impairment RESTART IDENTITY;"))
            df_ecl.to_sql("fact_ecl_impairment", conn, if_exists="append", index=False)

        print("🎉 IFRS 9 ECL results committed to PostgreSQL successfully!")

    def _export_excel_workbook(self, df, summary_df):
        """Generates an executive-formatted Excel workbook (.xlsx) using openpyxl."""
        excel_path = "reports/ifrs9_ecl_summary_model.xlsx"
        os.makedirs("reports", exist_ok=True)

        wb = Workbook()
        
        # Sheet 1: Staging Summary
        ws1 = wb.active
        ws1.title = "IFRS 9 Staging Summary"
        ws1.views.sheetView[0].showGridLines = True

        # Styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
        center_align = Alignment(horizontal="center", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")

        # Title Block
        ws1["A1"] = "IFRS 9 EXPECTED CREDIT LOSS (ECL) EXECUTIVE SUMMARY"
        ws1["A1"].font = title_font
        ws1["A2"] = f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Portfolio: Retail & SME Loans"
        ws1["A2"].font = Font(name="Calibri", size=10, italic=True, color="595959")

        # Write Summary Table Header
        headers_ws1 = ["IFRS 9 Stage", "Loan Count", "Total EAD (EUR)", "Total ECL Provision (EUR)", "Avg PD 12M (%)", "Avg LGD (%)", "Coverage Ratio (%)"]
        for col_num, header in enumerate(headers_ws1, 1):
            cell = ws1.cell(row=4, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # Write Summary Table Rows
        for r_idx, row in summary_df.iterrows():
            row_num = 5 + r_idx
            ws1.cell(row=row_num, column=1, value=f"Stage {int(row['ifrs9_stage'])}").alignment = center_align
            ws1.cell(row=row_num, column=2, value=int(row['loan_count'])).alignment = center_align
            
            c_ead = ws1.cell(row=row_num, column=3, value=float(row['total_ead']))
            c_ead.number_format = '€#,##0.00'
            c_ead.alignment = right_align
            
            c_ecl = ws1.cell(row=row_num, column=4, value=float(row['total_ecl']))
            c_ecl.number_format = '€#,##0.00'
            c_ecl.alignment = right_align

            c_pd = ws1.cell(row=row_num, column=5, value=float(row['avg_pd_12m']))
            c_pd.number_format = '0.00%'
            c_pd.alignment = right_align

            c_lgd = ws1.cell(row=row_num, column=6, value=float(row['avg_lgd']))
            c_lgd.number_format = '0.00%'
            c_lgd.alignment = right_align

            c_cov = ws1.cell(row=row_num, column=7, value=float(row['coverage_ratio_pct']) / 100.0)
            c_cov.number_format = '0.00%'
            c_cov.alignment = right_align

        # Sheet 2: Top Impaired Loans
        ws2 = wb.create_sheet(title="Top Impaired Exposures")
        ws2.views.sheetView[0].showGridLines = True
        
        ws2["A1"] = "TOP 50 IMPAIRED CREDIT EXPOSURES (STAGE 2 & STAGE 3)"
        ws2["A1"].font = title_font

        headers_ws2 = ["Loan Code", "Borrower Code", "IFRS 9 Stage", "Days Past Due", "Scorecard Points", "EAD (EUR)", "PD 12M (%)", "LGD (%)", "ECL Provision (EUR)"]
        for col_num, header in enumerate(headers_ws2, 1):
            cell = ws2.cell(row=3, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        top_impaired = df[df['ifrs9_stage'] > 1].sort_values(by='ecl_amount', ascending=False).head(50)

        # FIXED ROW INDEXING HERE:
        for row_num, (_, row) in enumerate(top_impaired.iterrows(), start=4):
            ws2.cell(row=row_num, column=1, value=str(row['loan_code'])).alignment = center_align
            ws2.cell(row=row_num, column=2, value=str(row['borrower_code'])).alignment = center_align
            ws2.cell(row=row_num, column=3, value=f"Stage {int(row['ifrs9_stage'])}").alignment = center_align
            ws2.cell(row=row_num, column=4, value=int(row['days_past_due'])).alignment = center_align
            ws2.cell(row=row_num, column=5, value=int(row['scorecard_points'])).alignment = center_align
            
            c_ead = ws2.cell(row=row_num, column=6, value=float(row['ead']))
            c_ead.number_format = '€#,##0.00'
            c_ead.alignment = right_align

            c_pd = ws2.cell(row=row_num, column=7, value=float(row['pd_12m']))
            c_pd.number_format = '0.00%'
            c_pd.alignment = right_align

            c_lgd = ws2.cell(row=row_num, column=8, value=float(row['lgd']))
            c_lgd.number_format = '0.00%'
            c_lgd.alignment = right_align

            c_ecl = ws2.cell(row=row_num, column=9, value=float(row['ecl_amount']))
            c_ecl.number_format = '€#,##0.00'
            c_ecl.alignment = right_align

        # Auto-adjust column widths
        for ws in [ws1, ws2]:
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = col[0].column_letter
                ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

        wb.save(excel_path)
        print(f"📄 Auditable Financial Excel Workbook saved to {excel_path}")

if __name__ == "__main__":
    ecl_engine = IFRS9ExpectedCreditLossEngine(collateral_haircut=0.20)
    ecl_engine.calculate_ifrs9_ecl()